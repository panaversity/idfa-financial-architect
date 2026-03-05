# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""IDFA compliance auditor — checks all four guardrails on an Excel model.

Usage:
    uv run idfa_audit.py <file>

Output: JSON compliance report to stdout.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

# Regex to detect coordinate references like A1, B14, $C$10, $D7, E$3
COORD_RE = re.compile(r"(?<![A-Za-z_])\$?[A-Z]{1,3}\$?\d+(?![A-Za-z_(\d])")

# Complex formula keywords that require LaTeX verification
COMPLEX_KEYWORDS = {"WACC", "NPV", "DCF", "IRR", "XNPV", "XIRR", "MIRR"}

# Numeric literal in formulas (detects hardcoded constants like *0.60, +100000)
# Excludes 0, 1, and small integers commonly used in structural formulas (1+growth)
HARDCODED_RE = re.compile(r"(?<=[*+\-/,(])\s*\d+\.?\d*(?!\s*[A-Za-z_])")


def audit(file: str) -> dict:
    path = Path(file)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {file}"}))
        sys.exit(2)

    wb = openpyxl.load_workbook(path)

    # Identify Calculations sheet (heuristic: sheet named "Calculations" or second sheet)
    calc_sheet_name = None
    if "Calculations" in wb.sheetnames:
        calc_sheet_name = "Calculations"
    elif len(wb.sheetnames) >= 2:
        calc_sheet_name = wb.sheetnames[1]

    if calc_sheet_name is None:
        print(json.dumps({"error": "No Calculations sheet found"}))
        sys.exit(2)

    ws = wb[calc_sheet_name]

    # Collect all formula cells in Calculations layer
    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append(cell)

    total_formulas = len(formula_cells)

    # --- Guardrail 1: Named Range Priority ---
    coord_violations = []
    for cell in formula_cells:
        formula = cell.value[1:]  # strip leading =
        matches = COORD_RE.findall(formula)
        if matches:
            coord_violations.append(
                {
                    "cell": cell.coordinate,
                    "formula": cell.value,
                    "coordinate_refs": matches,
                }
            )

    g1_status = "PASS" if len(coord_violations) == 0 else "FAIL"

    # --- Guardrail 2: LaTeX Verification ---
    complex_without_verification = []
    for cell in formula_cells:
        formula_upper = cell.value.upper()
        for kw in COMPLEX_KEYWORDS:
            if kw in formula_upper:
                comment = cell.comment
                has_latex = comment and (
                    "FORMULA:" in comment.text or "LaTeX" in comment.text.lower()
                )
                if not has_latex:
                    complex_without_verification.append(
                        {
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "keyword": kw,
                        }
                    )
                break

    g2_status = "PASS" if len(complex_without_verification) == 0 else "WARN"

    # --- Guardrail 3: Intent Notes ---
    formulas_with_comments = sum(1 for c in formula_cells if c.comment is not None)
    coverage = (formulas_with_comments / total_formulas * 100) if total_formulas > 0 else 100
    missing_notes = []
    for cell in formula_cells:
        if cell.comment is None:
            missing_notes.append({"cell": cell.coordinate, "formula": cell.value})

    g3_status = "PASS" if coverage == 100 else "FAIL"

    # --- Guardrail 4: Layer Isolation ---
    hardcoded_violations = []
    for cell in formula_cells:
        formula = cell.value[1:]  # strip leading =
        matches = HARDCODED_RE.findall(formula)
        # Filter out structural constants (0, 1) commonly used in formulas like (1+growth)
        significant = [m.strip() for m in matches if float(m.strip()) not in (0, 1)]
        if significant:
            hardcoded_violations.append(
                {
                    "cell": cell.coordinate,
                    "formula": cell.value,
                    "hardcoded_values": significant,
                }
            )

    g4_status = "PASS" if len(hardcoded_violations) == 0 else "FAIL"

    # --- Overall ---
    all_pass = all(s == "PASS" for s in [g1_status, g3_status, g4_status])
    total_violations = len(coord_violations) + len(hardcoded_violations) + len(missing_notes)
    total_checks = total_formulas * 3  # 3 checkable guardrails per formula
    score = ((total_checks - total_violations) / total_checks * 100) if total_checks > 0 else 100

    return {
        "file": file,
        "total_formulas": total_formulas,
        "guardrail_1_named_ranges": {
            "status": g1_status,
            "violations": len(coord_violations),
            "details": coord_violations,
        },
        "guardrail_2_latex": {
            "status": g2_status,
            "complex_formulas_without_verification": len(complex_without_verification),
            "details": complex_without_verification,
        },
        "guardrail_3_intent_notes": {
            "status": g3_status,
            "coverage": f"{coverage:.0f}%",
            "missing": missing_notes,
        },
        "guardrail_4_layer_isolation": {
            "status": g4_status,
            "violations": len(hardcoded_violations),
            "details": hardcoded_violations,
        },
        "overall": "PASS" if all_pass else "FAIL",
        "compliance_score": f"{score:.0f}%",
    }


def main() -> None:
    if len(sys.argv) != 2:
        print("Usage: idfa_audit.py <file>", file=sys.stderr)
        sys.exit(2)

    result = audit(sys.argv[1])
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
