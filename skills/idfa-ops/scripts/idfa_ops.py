# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""IDFA Named Range CRUD operations for Excel financial models.

Usage:
    uv run idfa_ops.py write <file> <name> <value>
    uv run idfa_ops.py read <file> <name> [name2 ...]
    uv run idfa_ops.py inspect <file>
    uv run idfa_ops.py formula <file> <name>
    uv run idfa_ops.py create-range <file> <name> <sheet> <cell>

All output is JSON to stdout. Exit codes: 0=success, 1=not found, 2=file error.
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName


def _resolve_named_range(wb: openpyxl.Workbook, name: str) -> tuple[str, str] | None:
    """Resolve a Named Range to (sheet_title, cell_coordinate) or None."""
    defn = wb.defined_names.get(name)
    if defn is None:
        return None
    # DefinedName.destinations yields (sheet_title, cell_range) tuples
    for sheet_title, cell_range in defn.destinations:
        # cell_range is like '$B$2' — strip $ signs
        cell = cell_range.replace("$", "")
        return (sheet_title, cell)
    return None


def _error(msg: str, code: int = 1) -> None:
    print(json.dumps({"error": msg}), file=sys.stdout)
    sys.exit(code)


def cmd_write(file: str, name: str, raw_value: str) -> None:
    path = Path(file)
    if not path.exists():
        _error(f"File not found: {file}", 2)

    wb = openpyxl.load_workbook(path)
    dest = _resolve_named_range(wb, name)
    if dest is None:
        _error(f"Named Range not found: {name}")

    sheet_title, cell = dest
    ws = wb[sheet_title]

    # Try to parse as number
    try:
        value = int(raw_value)
    except ValueError:
        try:
            value = float(raw_value)
        except ValueError:
            value = raw_value

    ws[cell].value = value
    wb.save(path)
    print(
        json.dumps({"status": "ok", "name": name, "value": value, "cell": f"{sheet_title}!{cell}"})
    )


def cmd_read(file: str, names: list[str]) -> None:
    path = Path(file)
    if not path.exists():
        _error(f"File not found: {file}", 2)

    wb = openpyxl.load_workbook(path, data_only=True)
    results = {}
    for name in names:
        dest = _resolve_named_range(wb, name)
        if dest is None:
            _error(f"Named Range not found: {name}")
        sheet_title, cell = dest
        ws = wb[sheet_title]
        results[name] = ws[cell].value

    print(json.dumps({"status": "ok", "values": results}))


def cmd_inspect(file: str) -> None:
    path = Path(file)
    if not path.exists():
        _error(f"File not found: {file}", 2)

    wb_formulas = openpyxl.load_workbook(path)
    wb_values = openpyxl.load_workbook(path, data_only=True)

    ranges = []
    for name in wb_formulas.defined_names:
        defn = wb_formulas.defined_names[name]
        entry: dict = {"name": defn.name}
        try:
            for sheet_title, cell_range in defn.destinations:
                cell = cell_range.replace("$", "")
                entry["sheet"] = sheet_title
                entry["cell"] = cell

                ws_f = wb_formulas[sheet_title]
                cell_obj_f = ws_f[cell]
                formula = cell_obj_f.value
                entry["formula"] = (
                    formula if isinstance(formula, str) and formula.startswith("=") else None
                )
                entry["value_raw"] = formula if entry["formula"] is None else None

                ws_v = wb_values[sheet_title]
                cell_obj_v = ws_v[cell]
                entry["cached_value"] = cell_obj_v.value

                comment = cell_obj_f.comment
                entry["intent_note"] = comment.text if comment else None
                break
        except Exception as e:
            entry["error"] = str(e)

        ranges.append(entry)

    print(json.dumps({"status": "ok", "named_ranges": ranges, "count": len(ranges)}))


def cmd_formula(file: str, name: str) -> None:
    path = Path(file)
    if not path.exists():
        _error(f"File not found: {file}", 2)

    wb = openpyxl.load_workbook(path)  # NOT data_only — get formulas
    dest = _resolve_named_range(wb, name)
    if dest is None:
        _error(f"Named Range not found: {name}")

    sheet_title, cell = dest
    ws = wb[sheet_title]
    value = ws[cell].value

    if isinstance(value, str) and value.startswith("="):
        print(json.dumps({"status": "ok", "name": name, "formula": value}))
    else:
        print(json.dumps({"status": "ok", "name": name, "formula": None, "value": value}))


def cmd_create_range(file: str, name: str, sheet: str, cell: str) -> None:
    path = Path(file)
    if not path.exists():
        _error(f"File not found: {file}", 2)

    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        _error(f"Sheet not found: {sheet}")

    ref = f"'{sheet}'!${cell[0]}${cell[1:]}"
    defn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(defn)
    wb.save(path)
    print(json.dumps({"status": "ok", "name": name, "reference": f"{sheet}!{cell}"}))


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: idfa_ops.py <command> <args...>", file=sys.stderr)
        print("Commands: write, read, inspect, formula, create-range", file=sys.stderr)
        sys.exit(2)

    cmd = sys.argv[1]

    if cmd == "write":
        if len(sys.argv) != 5:
            _error("Usage: write <file> <name> <value>", 2)
        cmd_write(sys.argv[2], sys.argv[3], sys.argv[4])

    elif cmd == "read":
        if len(sys.argv) < 4:
            _error("Usage: read <file> <name> [name2 ...]", 2)
        cmd_read(sys.argv[2], sys.argv[3:])

    elif cmd == "inspect":
        if len(sys.argv) != 3:
            _error("Usage: inspect <file>", 2)
        cmd_inspect(sys.argv[2])

    elif cmd == "formula":
        if len(sys.argv) != 4:
            _error("Usage: formula <file> <name>", 2)
        cmd_formula(sys.argv[2], sys.argv[3])

    elif cmd == "create-range":
        if len(sys.argv) != 6:
            _error("Usage: create-range <file> <name> <sheet> <cell>", 2)
        cmd_create_range(sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5])

    else:
        _error(f"Unknown command: {cmd}", 2)


if __name__ == "__main__":
    main()
