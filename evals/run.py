# /// script
# requires-python = ">=3.10"
# dependencies = ["pyyaml>=6.0", "openpyxl>=3.1"]
# ///
"""IDFA Eval Harness — two-tier evaluation of the Financial Architect skill.

Tier 1: Deterministic checks (idfa_audit.py, artifact deltas, row insertion)
Tier 2: LLM judges via `claude -p` headless (decision_ready, risk_awareness, etc.)

Usage:
    uv run evals/run.py                          # full eval, all cases
    uv run evals/run.py --case build_gp_waterfall # single case
    uv run evals/run.py --runs 3                  # stability check
    uv run evals/run.py --calibrate               # judge calibration
    uv run evals/run.py --verbose                  # real-time output
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import yaml

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent  # repo root
EVALS_DIR = Path(__file__).resolve().parent
SCRIPTS_DIR = ROOT / "skills" / "idfa-ops" / "scripts"
GRADERS_DIR = EVALS_DIR / "graders"
RESULTS_DIR = EVALS_DIR / "results"
GOLDEN_DIR = EVALS_DIR / "golden"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def log(msg: str, verbose: bool = True) -> None:
    if verbose:
        print(f"  {msg}", file=sys.stderr)


def run_cmd(cmd: list[str], timeout: int = 300) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)


def load_cases(path: Path) -> list[dict]:
    with open(path) as f:
        data = yaml.safe_load(f)
    return data["cases"]


# ---------------------------------------------------------------------------
# Tier 1 — Deterministic checks
# ---------------------------------------------------------------------------


def run_audit(xlsx_path: str) -> dict:
    """Run idfa_audit.py on an xlsx file, return parsed JSON."""
    result = run_cmd(["uv", "run", str(SCRIPTS_DIR / "idfa_audit.py"), xlsx_path])
    if result.returncode != 0:
        return {"error": f"audit failed: {result.stderr.strip()}"}
    return json.loads(result.stdout)


def snapshot_named_ranges(xlsx_path: str) -> dict:
    """Read all named range values from an xlsx (data_only mode)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    values = {}
    for name in wb.defined_names:
        defn = wb.defined_names[name]
        try:
            for sheet_title, cell_range in defn.destinations:
                cell = cell_range.replace("$", "")
                ws = wb[sheet_title]
                values[defn.name] = ws[cell].value
                break
        except Exception:
            values[defn.name] = None
    return values


def check_artifact_delta(before: dict, after: dict, expected_delta: dict) -> dict:
    """Verify expected named range changes occurred."""
    failures = []
    for range_name, expected_val in expected_delta.items():
        before_val = before.get(range_name)
        after_val = after.get(range_name)
        if after_val is None:
            failures.append(f"{range_name}: not found in output")
        elif isinstance(expected_val, (int, float)) and isinstance(after_val, (int, float)):
            if abs(after_val - expected_val) > 0.01 * abs(expected_val):
                failures.append(f"{range_name}: expected ~{expected_val}, got {after_val}")
        elif after_val != expected_val:
            failures.append(f"{range_name}: expected {expected_val}, got {after_val}")

        # Verify it actually changed from before
        if before_val == after_val and before_val is not None:
            failures.append(f"{range_name}: value unchanged (still {before_val})")

    return {"pass": len(failures) == 0, "failures": failures}


def _try_recalc(xlsx_path: str) -> bool:
    """Attempt to recalculate the xlsx via recalc_bridge. Returns True on success."""
    recalc_script = SCRIPTS_DIR / "recalc_bridge.py"
    if not recalc_script.exists():
        return False
    try:
        result = run_cmd(["uv", "run", str(recalc_script), xlsx_path], timeout=120)
        if result.returncode == 0:
            output = json.loads(result.stdout)
            return output.get("status") == "ok"
    except Exception:
        pass
    return False


def check_target_reached(
    xlsx_path: str,
    range_name: str,
    target: float,
    tolerance: float,
    solve_input: dict | None = None,
) -> dict:
    """Verify a named range value is within tolerance of target.

    If the target range has no cached value (openpyxl can't recalculate),
    attempts recalc_bridge first.  As a final fallback, if ``solve_input``
    is provided, checks that the correct *input* Named Range was written —
    proving the agent found the right answer even without a recalc engine.
    """
    # Attempt recalc before reading (populates cached values if engine exists)
    _try_recalc(xlsx_path)

    values = snapshot_named_ranges(xlsx_path)
    actual = values.get(range_name)

    # Happy path: cached value exists (recalc engine available)
    if actual is not None and isinstance(actual, (int, float)):
        pct_off = abs(actual - target) / abs(target) if target != 0 else abs(actual)
        passed = pct_off <= tolerance
        return {
            "pass": passed,
            "actual": actual,
            "target": target,
            "pct_off": f"{pct_off:.1%}",
            "tolerance": f"{tolerance:.0%}",
        }

    # Fallback: no cached value — check the input Named Range instead.
    # If the agent wrote the correct input, the model WOULD produce the target
    # once recalculated (formula correctness verified by g1_named_ranges).
    if solve_input:
        input_name = solve_input["range_name"]
        expected_input = solve_input["value"]
        input_tolerance = solve_input.get("tolerance", tolerance)
        actual_input = values.get(input_name)
        if actual_input is not None and isinstance(actual_input, (int, float)):
            pct_off = (
                abs(actual_input - expected_input) / abs(expected_input)
                if expected_input != 0
                else abs(actual_input)
            )
            passed = pct_off <= input_tolerance
            return {
                "pass": passed,
                "method": "solve_input_fallback",
                "input_name": input_name,
                "actual_input": actual_input,
                "expected_input": expected_input,
                "pct_off": f"{pct_off:.1%}",
                "tolerance": f"{input_tolerance:.0%}",
                "note": f"{range_name} cached value unavailable (no recalc engine); verified via input",
            }

    return {
        "pass": False,
        "actual": None,
        "reason": f"{range_name} not found (no cached value, no recalc engine)",
    }


def check_layer_separation(xlsx_path: str) -> dict:
    """Verify Inp_* ranges -> Assumptions sheet, calc ranges -> Calculations sheet."""
    wb = openpyxl.load_workbook(xlsx_path)
    failures = []
    for name in wb.defined_names:
        defn = wb.defined_names[name]
        try:
            for sheet_title, _ in defn.destinations:
                if defn.name.startswith("Inp_") and sheet_title != "Assumptions":
                    failures.append(f"{defn.name} on {sheet_title}, expected Assumptions")
                elif not defn.name.startswith("Inp_") and sheet_title == "Assumptions":
                    # Non-input ranges should not be on Assumptions
                    pass  # Some models may have labels; only flag calc ranges
                break
        except Exception:
            pass
    return {"pass": len(failures) == 0, "failures": failures}


def check_row_insertion(xlsx_path: str) -> dict:
    """Insert a row into Calculations, verify Named Range outputs unchanged."""
    # Snapshot before
    before = snapshot_named_ranges(xlsx_path)

    # Copy to temp
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    shutil.copy2(xlsx_path, tmp.name)

    try:
        wb = openpyxl.load_workbook(tmp.name)
        if "Calculations" not in wb.sheetnames:
            return {"pass": False, "reason": "No Calculations sheet"}

        ws = wb["Calculations"]
        # Insert a row in the middle of the calculations block
        ws.insert_rows(5)
        wb.save(tmp.name)

        # Snapshot after — named ranges should still resolve correctly
        # (openpyxl adjusts cell references in named ranges on row insert)
        after = snapshot_named_ranges(tmp.name)

        # Compare: all named ranges that had values should still have them
        mismatches = []
        for name, val in before.items():
            if name not in after:
                mismatches.append(f"{name}: missing after row insertion")
            elif after[name] != val and val is not None:
                mismatches.append(f"{name}: was {val}, now {after[name]}")

        return {"pass": len(mismatches) == 0, "mismatches": mismatches}
    finally:
        os.unlink(tmp.name)


def check_audit_claims(
    agent_output: str,
    audit_result: dict,
    expected_result: str,
    expected_violations: list[str] | None = None,
) -> dict:
    """For audit cases: compare agent's narrative claims against deterministic audit."""
    failures = []
    output_lower = agent_output.lower()

    if expected_result == "PASS":
        # Agent should say it's compliant
        compliant_indicators = ["compliant", "pass", "all clear", "no violations", "meets"]
        if not any(ind in output_lower for ind in compliant_indicators):
            failures.append("Agent did not indicate model is compliant, but audit says PASS")
    elif expected_result == "FAIL":
        # Agent should identify violations
        if expected_violations:
            violation_keywords = {
                "guardrail_1_named_ranges": [
                    "coordinate",
                    "cell reference",
                    "named range",
                    "b2",
                    "b3",
                ],
                "guardrail_3_intent_notes": [
                    "intent note",
                    "comment",
                    "missing note",
                    "documentation",
                ],
                "guardrail_4_layer_isolation": ["hardcoded", "constant", "0.60", "layer isolation"],
            }
            for v in expected_violations:
                keywords = violation_keywords.get(v, [])
                if keywords and not any(kw in output_lower for kw in keywords):
                    failures.append(f"Agent missed violation: {v}")

    return {"pass": len(failures) == 0, "failures": failures}


def run_tier1(case: dict, agent_output: str, work_dir: str, verbose: bool = False) -> dict:
    """Execute all Tier 1 checks for a case. Returns dict of check_name -> result."""
    results = {}
    checks = case.get("tier1_checks", [])

    # Find the output xlsx (agent may have created it or modified fixture in-place)
    output_xlsx = None
    if case.get("fixture"):
        # For cases with fixtures, the working copy is in work_dir
        fixture_name = Path(case["fixture"]).name
        candidate = os.path.join(work_dir, fixture_name)
        if os.path.exists(candidate):
            output_xlsx = candidate
        else:
            # Search for any xlsx in work_dir
            for f in os.listdir(work_dir):
                if f.endswith(".xlsx"):
                    output_xlsx = os.path.join(work_dir, f)
                    break
        # Fallback: agent may have modified the original fixture in the repo
        if output_xlsx is None:
            repo_fixture = str(ROOT / case["fixture"])
            if os.path.exists(repo_fixture):
                output_xlsx = repo_fixture
                log(
                    "Warning: using repo fixture as fallback — agent may have modified it in-place",
                    verbose,
                )
    else:
        # Build case — find any xlsx the agent created
        for f in os.listdir(work_dir):
            if f.endswith(".xlsx"):
                output_xlsx = os.path.join(work_dir, f)
                break

    # Run audit if any guardrail checks needed
    # Skip guardrail checks for audit operations — the fixture may intentionally have
    # violations. For audit cases, we check agent claims via audit_claims instead.
    audit_checks = {"g1_named_ranges", "g3_intent_notes", "g4_layer_isolation"}
    if audit_checks.intersection(checks) and output_xlsx and case.get("operation") != "audit":
        log(f"Running audit on {Path(output_xlsx).name}", verbose)
        audit = run_audit(output_xlsx)
        if "error" in audit:
            for c in audit_checks.intersection(checks):
                results[c] = {"pass": False, "error": audit["error"]}
        else:
            if "g1_named_ranges" in checks:
                results["g1_named_ranges"] = {
                    "pass": audit["guardrail_1_named_ranges"]["status"] == "PASS",
                    "details": audit["guardrail_1_named_ranges"],
                }
            if "g3_intent_notes" in checks:
                results["g3_intent_notes"] = {
                    "pass": audit["guardrail_3_intent_notes"]["status"] == "PASS",
                    "details": audit["guardrail_3_intent_notes"],
                }
            if "g4_layer_isolation" in checks:
                results["g4_layer_isolation"] = {
                    "pass": audit["guardrail_4_layer_isolation"]["status"] == "PASS",
                    "details": audit["guardrail_4_layer_isolation"],
                }

    # Audit claim check (for audit operation cases)
    if case.get("operation") == "audit" and output_xlsx:
        fixture_path = str(ROOT / case["fixture"])
        audit = run_audit(fixture_path)
        claim_result = check_audit_claims(
            agent_output,
            audit,
            case.get("expected_audit_result", "PASS"),
            case.get("expected_violations"),
        )
        results["audit_claims"] = claim_result

    # Retrofit check: before should FAIL, after should PASS
    if case.get("operation") == "retrofit" and output_xlsx:
        fixture_path = str(ROOT / case["fixture"])
        before_audit = run_audit(fixture_path)
        after_audit = run_audit(output_xlsx)
        before_ok = before_audit.get("overall") == case.get("expected_before_audit", "FAIL")
        after_ok = after_audit.get("overall") == case.get("expected_after_audit", "PASS")
        results["retrofit_delta"] = {
            "pass": before_ok and after_ok,
            "before": before_audit.get("overall"),
            "after": after_audit.get("overall"),
        }

    # Artifact delta
    if "artifact_delta" in checks and output_xlsx and case.get("expected_delta"):
        fixture_path = str(ROOT / case["fixture"])
        before = snapshot_named_ranges(fixture_path)
        after = snapshot_named_ranges(output_xlsx)
        results["artifact_delta"] = check_artifact_delta(before, after, case["expected_delta"])
        log(f"Artifact delta: {results['artifact_delta']}", verbose)

    # Target reached
    if "target_reached" in checks and case.get("target"):
        t = case["target"]
        if output_xlsx:
            results["target_reached"] = check_target_reached(
                output_xlsx,
                t["range_name"],
                t["value"],
                t["tolerance"],
                solve_input=t.get("solve_input"),
            )
        else:
            results["target_reached"] = {
                "pass": False,
                "actual": None,
                "reason": "No output xlsx found — agent may have answered textually without saving",
            }
        log(f"Target reached: {results['target_reached']}", verbose)

    # Layer separation
    if "layer_separation" in checks and output_xlsx:
        results["layer_separation"] = check_layer_separation(output_xlsx)

    # Row insertion robustness
    if "row_insertion" in checks:
        fixture_path = str(ROOT / case["fixture"])
        results["row_insertion"] = check_row_insertion(fixture_path)
        log(f"Row insertion: {results['row_insertion']}", verbose)

    return results


# ---------------------------------------------------------------------------
# Tier 2 — LLM judges
# ---------------------------------------------------------------------------


def _extract_structured_output(response: dict | list) -> dict:
    """Extract structured output from claude -p --output-format json response.

    The response is a JSON array of conversation events. The structured output
    lives inside the assistant message's StructuredOutput tool_use content block.
    """
    # If it's already a dict with our expected fields, return directly
    if isinstance(response, dict):
        if "pass" in response:
            return response
        if "structured_output" in response:
            return response["structured_output"]

    # JSON array of conversation events
    if isinstance(response, list):
        for event in response:
            if not isinstance(event, dict):
                continue
            # Look for assistant message with StructuredOutput tool use
            if event.get("type") == "assistant":
                msg = event.get("message", {})
                for block in msg.get("content", []):
                    if (
                        isinstance(block, dict)
                        and block.get("name") == "StructuredOutput"
                        and "input" in block
                    ):
                        return block["input"]

    return {"pass": False, "reasoning": "Could not extract structured output from response"}


def run_judge(grader_name: str, agent_output: str, operation: str) -> dict:
    """Run a single LLM judge via `claude -p` with JSON schema enforcement."""
    grader_path = GRADERS_DIR / f"{grader_name}.txt"
    if not grader_path.exists():
        return {"pass": False, "reasoning": f"Grader file not found: {grader_path}"}

    grader_prompt = grader_path.read_text()

    prompt = f"""{grader_prompt}

---

OPERATION TYPE: {operation}

AGENT OUTPUT TO EVALUATE:
{agent_output}
"""

    schema = json.dumps(
        {
            "type": "object",
            "properties": {"pass": {"type": "boolean"}, "reasoning": {"type": "string"}},
            "required": ["pass", "reasoning"],
        }
    )

    result = run_cmd(
        [
            "claude",
            "-p",
            prompt,
            "--output-format",
            "json",
            "--json-schema",
            schema,
            "--max-turns",
            "2",
            "--model",
            "claude-haiku-4-5-20251001",
        ],
        timeout=120,
    )

    if result.returncode != 0:
        return {"pass": False, "reasoning": f"claude CLI error: {result.stderr.strip()[:200]}"}

    try:
        response = json.loads(result.stdout)
        return _extract_structured_output(response)
    except json.JSONDecodeError:
        return {
            "pass": False,
            "reasoning": f"Failed to parse judge response: {result.stdout[:200]}",
        }


def run_tier2(case: dict, agent_output: str, verbose: bool = False) -> dict:
    """Execute all Tier 2 LLM judges in parallel."""
    judges = case.get("tier2_judges", [])
    if not judges:
        return {}

    results = {}
    operation = case["operation"]

    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {
            pool.submit(run_judge, judge, agent_output, operation): judge for judge in judges
        }
        for future in as_completed(futures):
            judge_name = futures[future]
            try:
                result = future.result()
                results[judge_name] = result
                log(f"Judge {judge_name}: {'PASS' if result.get('pass') else 'FAIL'}", verbose)
            except Exception as e:
                results[judge_name] = {"pass": False, "reasoning": f"Exception: {e}"}

    return results


# ---------------------------------------------------------------------------
# Agent runner
# ---------------------------------------------------------------------------


def run_agent(case: dict, work_dir: str, verbose: bool = False) -> str:
    """Run the IDFA Financial Architect skill via `claude -p`."""
    prompt = case["prompt"]

    # If fixture, reference the copy in work_dir
    if case.get("fixture"):
        fixture_name = Path(case["fixture"]).name
        fixture_in_workdir = os.path.join(work_dir, fixture_name)
        prompt += f"\n\nThe model file is at: {fixture_in_workdir}"

    prompt += f"\n\nWorking directory for any output files: {work_dir}"

    log(f"Running agent for {case['id']}...", verbose)
    result = run_cmd(
        [
            "claude",
            "-p",
            prompt,
            "--allowedTools",
            "mcp__idfa-ops__*,Write,Read,Bash",
            "--plugin-dir",
            str(ROOT),
            "--output-format",
            "text",
            "--max-turns",
            "20",
        ],
        timeout=600,
    )

    if result.returncode != 0:
        return f"[AGENT ERROR] exit={result.returncode}\nstderr: {result.stderr.strip()[:500]}"

    return result.stdout.strip()


# ---------------------------------------------------------------------------
# Single case runner
# ---------------------------------------------------------------------------


def run_case(case: dict, verbose: bool = False) -> dict:
    """Run a single eval case: agent -> tier1 -> tier2 -> aggregate."""
    case_id = case["id"]
    operation = case["operation"]

    log(f"\n{'=' * 60}", verbose)
    log(f"Case: {case_id} ({operation})", verbose)
    log(f"{'=' * 60}", verbose)

    # Create temp work dir
    work_dir = tempfile.mkdtemp(prefix=f"idfa_eval_{case_id}_")

    # Copy fixture if needed
    if case.get("fixture"):
        src = ROOT / case["fixture"]
        dst = os.path.join(work_dir, src.name)
        shutil.copy2(str(src), dst)

    agent_output = ""
    tier1_results = {}
    tier2_results = {}

    # Programmatic cases (no agent run)
    if operation == "programmatic":
        log("Programmatic test — no agent run", verbose)
        tier1_results = run_tier1(case, "", work_dir, verbose)
    else:
        # Run agent
        agent_output = run_agent(case, work_dir, verbose)
        if verbose:
            preview = agent_output[:300] + ("..." if len(agent_output) > 300 else "")
            log(f"Agent output preview: {preview}", verbose)

        # Tier 1 — deterministic (runs first, fast)
        tier1_results = run_tier1(case, agent_output, work_dir, verbose)

        # Tier 2 — LLM judges (always run for full data)
        tier2_results = run_tier2(case, agent_output, verbose)

    # Aggregate
    all_tier1_pass = (
        all(r.get("pass", False) for r in tier1_results.values()) if tier1_results else True
    )
    all_tier2_pass = (
        all(r.get("pass", False) for r in tier2_results.values()) if tier2_results else True
    )
    case_pass = all_tier1_pass and all_tier2_pass

    # Cleanup work dir
    try:
        shutil.rmtree(work_dir)
    except Exception:
        pass

    return {
        "id": case_id,
        "operation": operation,
        "pass": case_pass,
        "agent_output_length": len(agent_output),
        "tier1": tier1_results,
        "tier2": tier2_results,
    }


# ---------------------------------------------------------------------------
# Calibration
# ---------------------------------------------------------------------------


def run_calibration(verbose: bool = False) -> dict:
    """Score golden set and report per-judge agreement rate."""
    if not GOLDEN_DIR.exists():
        print("No golden/ directory found. Create labeled examples first.")
        sys.exit(1)

    results = {}
    golden_files = sorted(GOLDEN_DIR.glob("*.txt"))

    if not golden_files:
        print("No golden set files found in evals/golden/")
        sys.exit(1)

    # Parse filenames: {grader}_{pass|fail}_{n}.txt
    for f in golden_files:
        parts = f.stem.rsplit("_", 2)
        if len(parts) < 3:
            log(f"Skipping {f.name} — expected format: grader_pass/fail_N.txt", verbose)
            continue

        grader = parts[0]
        expected = parts[1] == "pass"
        content = f.read_text()

        # Extract operation from first line if present, default to "general"
        lines = content.strip().split("\n")
        operation = "general"
        if lines[0].startswith("OPERATION:"):
            operation = lines[0].split(":", 1)[1].strip()
            content = "\n".join(lines[1:])

        if grader not in results:
            results[grader] = {"total": 0, "agree": 0, "details": []}

        judge_result = run_judge(grader, content, operation)
        actual = judge_result.get("pass", False)
        agreed = actual == expected

        results[grader]["total"] += 1
        if agreed:
            results[grader]["agree"] += 1
        results[grader]["details"].append(
            {
                "file": f.name,
                "expected": expected,
                "actual": actual,
                "agreed": agreed,
                "reasoning": judge_result.get("reasoning", ""),
            }
        )

        log(
            f"  {f.name}: expected={'PASS' if expected else 'FAIL'}, "
            f"got={'PASS' if actual else 'FAIL'} "
            f"{'OK' if agreed else 'MISMATCH'}",
            verbose,
        )

    return results


# ---------------------------------------------------------------------------
# Console report
# ---------------------------------------------------------------------------


def print_report(
    all_results: list[dict], run_count: int = 1, stability: dict | None = None
) -> None:
    """Print formatted console report."""
    ts = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    total = len(all_results)
    passed = sum(1 for r in all_results if r["pass"])

    print(f"\nIDFA Eval Results — {ts}")
    print("=" * 50)
    print(f"Overall: {passed}/{total} cases pass ({passed / total * 100:.0f}%)")

    # By operation
    ops = {}
    for r in all_results:
        op = r["operation"]
        if op not in ops:
            ops[op] = {"total": 0, "pass": 0}
        ops[op]["total"] += 1
        if r["pass"]:
            ops[op]["pass"] += 1

    print("\nBy Operation:")
    for op, counts in sorted(ops.items()):
        pct = counts["pass"] / counts["total"] * 100
        bar_len = int(pct / 100 * 12)
        bar = "\u2588" * bar_len + "\u2591" * (12 - bar_len)
        flag = " <-- NEEDS ATTENTION" if pct < 100 else ""
        print(f"  {op:<15} {counts['pass']}/{counts['total']}  {bar} {pct:3.0f}%{flag}")

    # Tier 1 aggregate
    tier1_agg = {}
    for r in all_results:
        for check_name, check_result in r.get("tier1", {}).items():
            if check_name not in tier1_agg:
                tier1_agg[check_name] = {"total": 0, "pass": 0}
            tier1_agg[check_name]["total"] += 1
            if check_result.get("pass", False):
                tier1_agg[check_name]["pass"] += 1

    if tier1_agg:
        print("\nDeterministic Checks (Tier 1):")
        for name, counts in sorted(tier1_agg.items()):
            pct = counts["pass"] / counts["total"] * 100
            status = "PASS" if pct == 100 else "FAIL" if pct == 0 else "PARTIAL"
            print(f"  {name:<25} {counts['pass']}/{counts['total']} {status:>7}  {pct:3.0f}%")

    # Tier 2 aggregate
    tier2_agg = {}
    for r in all_results:
        for judge_name, judge_result in r.get("tier2", {}).items():
            if judge_name not in tier2_agg:
                tier2_agg[judge_name] = {"total": 0, "pass": 0}
            tier2_agg[judge_name]["total"] += 1
            if judge_result.get("pass", False):
                tier2_agg[judge_name]["pass"] += 1

    if tier2_agg:
        print("\nLLM Judges (Tier 2):")
        for name, counts in sorted(tier2_agg.items()):
            pct = counts["pass"] / counts["total"] * 100
            print(f"  {name:<20} {counts['pass']}/{counts['total']}  {pct:3.0f}%")

    # Failures
    failures = [r for r in all_results if not r["pass"]]
    if failures:
        print("\nFailures:")
        for r in failures:
            print(f"  {r['id']}:")
            for check, result in r.get("tier1", {}).items():
                if not result.get("pass", False):
                    detail = result.get("failures", result.get("reason", result.get("details", "")))
                    print(f"    Tier 1 FAIL: {check} — {detail}")
            for judge, result in r.get("tier2", {}).items():
                if not result.get("pass", False):
                    print(f"    Tier 2 FAIL: {judge} — {result.get('reasoning', '')[:100]}")

    # Stability report
    if stability:
        print(f"\nStability ({run_count} runs):")
        for judge, flips in sorted(stability.items()):
            flag = " UNSTABLE — tighten pass/fail definitions" if flips > 1 else ""
            print(f"  {judge:<20} {flips} flips{flag}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="IDFA Eval Harness")
    parser.add_argument("--case", help="Run a single case by ID")
    parser.add_argument("--verbose", "-v", action="store_true", help="Real-time output")
    parser.add_argument("--runs", type=int, default=1, help="Number of runs for stability check")
    parser.add_argument(
        "--calibrate", action="store_true", help="Score golden set for judge calibration"
    )
    args = parser.parse_args()

    # Calibration mode
    if args.calibrate:
        print("Running judge calibration on golden set...")
        results = run_calibration(args.verbose)
        print("\nCalibration Results:")
        print("=" * 50)
        for grader, data in sorted(results.items()):
            rate = data["agree"] / data["total"] * 100 if data["total"] > 0 else 0
            status = "OK" if rate >= 85 else "NEEDS WORK"
            print(f"  {grader:<20} {data['agree']}/{data['total']} agree ({rate:.0f}%) [{status}]")
            for d in data["details"]:
                if not d["agreed"]:
                    print(
                        f"    MISMATCH: {d['file']} (expected {'PASS' if d['expected'] else 'FAIL'}, "
                        f"got {'PASS' if d['actual'] else 'FAIL'})"
                    )
                    print(f"    Reasoning: {d['reasoning'][:120]}")
        return

    # Load cases
    cases = load_cases(EVALS_DIR / "cases.yaml")

    if args.case:
        cases = [c for c in cases if c["id"] == args.case]
        if not cases:
            print(f"Case '{args.case}' not found.", file=sys.stderr)
            sys.exit(1)

    # Ensure results dir
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    # Run
    all_run_results = []
    for run_idx in range(args.runs):
        if args.runs > 1:
            print(f"\n--- Run {run_idx + 1}/{args.runs} ---", file=sys.stderr)

        run_results = []
        for case in cases:
            result = run_case(case, args.verbose)
            run_results.append(result)

        all_run_results.append(run_results)

    # Use last run for primary report
    final_results = all_run_results[-1]

    # Stability analysis
    stability = None
    if args.runs > 1:
        stability = {}
        # Track per-judge flips
        for case_idx, case in enumerate(cases):
            for judge in case.get("tier2_judges", []):
                verdicts = []
                for run_results in all_run_results:
                    if case_idx < len(run_results):
                        t2 = run_results[case_idx].get("tier2", {})
                        if judge in t2:
                            verdicts.append(t2[judge].get("pass", False))
                if len(set(verdicts)) > 1:
                    key = f"{judge}/{case['id']}"
                    stability[key] = stability.get(key, 0) + 1

    print_report(final_results, args.runs, stability)

    # Write results JSON
    ts = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    output = {
        "timestamp": ts,
        "runs": args.runs,
        "cases": final_results,
    }
    if stability:
        output["stability"] = stability

    result_path = RESULTS_DIR / f"{ts}.json"
    with open(result_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    # Also write latest.json symlink/copy
    latest_path = RESULTS_DIR / "latest.json"
    with open(latest_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nResults written to: {result_path}")


if __name__ == "__main__":
    main()
