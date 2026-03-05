# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Generate IDFA test fixture Excel workbooks deterministically."""

from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def _add_named_range(wb: openpyxl.Workbook, name: str, sheet: str, cell: str) -> None:
    ref = f"'{sheet}'!${cell[0]}${cell[1:]}"
    defn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(defn)


def _set_cell(ws, cell: str, value, *, formula: bool = False, comment: str | None = None):
    c = ws[cell]
    if formula:
        c.value = value  # formula string
    else:
        c.value = value
    if comment:
        c.comment = Comment(comment, "IDFA Generator")


def create_compliant() -> Path:
    """Create an IDFA-compliant GP Waterfall workbook."""
    wb = openpyxl.Workbook()

    # --- Assumptions sheet ---
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a["A1"] = "Input"
    ws_a["B1"] = "Value"

    _set_cell(ws_a, "A2", "Year 1 Revenue")
    _set_cell(ws_a, "B2", 10_000_000, comment="INTENT: Base year revenue\nGENERATED: fixture")
    _add_named_range(wb, "Inp_Rev_Y1", "Assumptions", "B2")

    _set_cell(ws_a, "A3", "Revenue Growth Rate")
    _set_cell(ws_a, "B3", 0.10, comment="INTENT: Annual revenue growth rate\nGENERATED: fixture")
    _add_named_range(wb, "Inp_Rev_Growth", "Assumptions", "B3")

    _set_cell(ws_a, "A4", "Year 1 COGS %")
    _set_cell(
        ws_a, "B4", 0.60, comment="INTENT: Year 1 COGS as percentage of revenue\nGENERATED: fixture"
    )
    _add_named_range(wb, "Inp_COGS_Pct_Y1", "Assumptions", "B4")

    _set_cell(ws_a, "A5", "Annual COGS Efficiency Gain")
    _set_cell(
        ws_a, "B5", 0.01, comment="INTENT: Annual COGS efficiency improvement\nGENERATED: fixture"
    )
    _add_named_range(wb, "Inp_COGS_Efficiency", "Assumptions", "B5")

    # --- Calculations sheet ---
    ws_c = wb.create_sheet("Calculations")
    ws_c["A1"] = "Variable"
    ws_c["B1"] = "Formula"

    # Revenue
    _set_cell(ws_c, "A2", "Revenue_Y1")
    _set_cell(
        ws_c,
        "B2",
        "=Inp_Rev_Y1",
        formula=True,
        comment="INTENT: Year 1 revenue equals the input assumption\nFORMULA: R_1 = Inp\nASSUMPTIONS: Inp_Rev_Y1",
    )
    _add_named_range(wb, "Revenue_Y1", "Calculations", "B2")

    _set_cell(ws_c, "A3", "Revenue_Y2")
    _set_cell(
        ws_c,
        "B3",
        "=Revenue_Y1*(1+Inp_Rev_Growth)",
        formula=True,
        comment="INTENT: Year 2 revenue = prior year * (1 + growth)\nFORMULA: R_2 = R_1 * (1+g)\nASSUMPTIONS: Revenue_Y1, Inp_Rev_Growth",
    )
    _add_named_range(wb, "Revenue_Y2", "Calculations", "B3")

    _set_cell(ws_c, "A4", "Revenue_Y3")
    _set_cell(
        ws_c,
        "B4",
        "=Revenue_Y2*(1+Inp_Rev_Growth)",
        formula=True,
        comment="INTENT: Year 3 revenue = prior year * (1 + growth)\nFORMULA: R_3 = R_2 * (1+g)\nASSUMPTIONS: Revenue_Y2, Inp_Rev_Growth",
    )
    _add_named_range(wb, "Revenue_Y3", "Calculations", "B4")

    # COGS %
    _set_cell(ws_c, "A5", "COGS_Pct_Y1")
    _set_cell(
        ws_c,
        "B5",
        "=Inp_COGS_Pct_Y1",
        formula=True,
        comment="INTENT: Year 1 COGS % equals the input assumption\nASSUMPTIONS: Inp_COGS_Pct_Y1",
    )
    _add_named_range(wb, "COGS_Pct_Y1", "Calculations", "B5")

    _set_cell(ws_c, "A6", "COGS_Pct_Y2")
    _set_cell(
        ws_c,
        "B6",
        "=COGS_Pct_Y1-Inp_COGS_Efficiency",
        formula=True,
        comment="INTENT: Year 2 COGS % = prior year - efficiency gain\nFORMULA: COGS%_2 = COGS%_1 - e\nASSUMPTIONS: COGS_Pct_Y1, Inp_COGS_Efficiency",
    )
    _add_named_range(wb, "COGS_Pct_Y2", "Calculations", "B6")

    _set_cell(ws_c, "A7", "COGS_Pct_Y3")
    _set_cell(
        ws_c,
        "B7",
        "=COGS_Pct_Y2-Inp_COGS_Efficiency",
        formula=True,
        comment="INTENT: Year 3 COGS % = prior year - efficiency gain\nFORMULA: COGS%_3 = COGS%_2 - e\nASSUMPTIONS: COGS_Pct_Y2, Inp_COGS_Efficiency",
    )
    _add_named_range(wb, "COGS_Pct_Y3", "Calculations", "B7")

    # COGS $
    _set_cell(ws_c, "A8", "COGS_Y1")
    _set_cell(
        ws_c,
        "B8",
        "=Revenue_Y1*COGS_Pct_Y1",
        formula=True,
        comment="INTENT: Year 1 COGS = revenue * COGS %\nFORMULA: COGS_1 = R_1 * COGS%_1\nASSUMPTIONS: Revenue_Y1, COGS_Pct_Y1",
    )
    _add_named_range(wb, "COGS_Y1", "Calculations", "B8")

    _set_cell(ws_c, "A9", "COGS_Y2")
    _set_cell(
        ws_c,
        "B9",
        "=Revenue_Y2*COGS_Pct_Y2",
        formula=True,
        comment="INTENT: Year 2 COGS = revenue * COGS %\nFORMULA: COGS_2 = R_2 * COGS%_2\nASSUMPTIONS: Revenue_Y2, COGS_Pct_Y2",
    )
    _add_named_range(wb, "COGS_Y2", "Calculations", "B9")

    _set_cell(ws_c, "A10", "COGS_Y3")
    _set_cell(
        ws_c,
        "B10",
        "=Revenue_Y3*COGS_Pct_Y3",
        formula=True,
        comment="INTENT: Year 3 COGS = revenue * COGS %\nFORMULA: COGS_3 = R_3 * COGS%_3\nASSUMPTIONS: Revenue_Y3, COGS_Pct_Y3",
    )
    _add_named_range(wb, "COGS_Y3", "Calculations", "B10")

    # Gross Profit
    _set_cell(ws_c, "A11", "Gross_Profit_Y1")
    _set_cell(
        ws_c,
        "B11",
        "=Revenue_Y1-COGS_Y1",
        formula=True,
        comment="INTENT: Year 1 Gross Profit = revenue - COGS\nFORMULA: GP_1 = R_1 - COGS_1\nASSUMPTIONS: Revenue_Y1, COGS_Y1",
    )
    _add_named_range(wb, "Gross_Profit_Y1", "Calculations", "B11")

    _set_cell(ws_c, "A12", "Gross_Profit_Y2")
    _set_cell(
        ws_c,
        "B12",
        "=Revenue_Y2-COGS_Y2",
        formula=True,
        comment="INTENT: Year 2 Gross Profit = revenue - COGS\nFORMULA: GP_2 = R_2 - COGS_2\nASSUMPTIONS: Revenue_Y2, COGS_Y2",
    )
    _add_named_range(wb, "Gross_Profit_Y2", "Calculations", "B12")

    _set_cell(ws_c, "A13", "Gross_Profit_Y3")
    _set_cell(
        ws_c,
        "B13",
        "=Revenue_Y3-COGS_Y3",
        formula=True,
        comment="INTENT: Year 3 Gross Profit = revenue - COGS\nFORMULA: GP_3 = R_3 - COGS_3\nASSUMPTIONS: Revenue_Y3, COGS_Y3",
    )
    _add_named_range(wb, "Gross_Profit_Y3", "Calculations", "B13")

    # --- Output sheet ---
    ws_o = wb.create_sheet("Output")
    ws_o["A1"] = "Item"
    ws_o["B1"] = "Year 1"
    ws_o["C1"] = "Year 2"
    ws_o["D1"] = "Year 3"

    ws_o["A2"] = "Revenue"
    ws_o["B2"] = "=Revenue_Y1"
    ws_o["C2"] = "=Revenue_Y2"
    ws_o["D2"] = "=Revenue_Y3"

    ws_o["A3"] = "COGS"
    ws_o["B3"] = "=COGS_Y1"
    ws_o["C3"] = "=COGS_Y2"
    ws_o["D3"] = "=COGS_Y3"

    ws_o["A4"] = "Gross Profit"
    ws_o["B4"] = "=Gross_Profit_Y1"
    ws_o["C4"] = "=Gross_Profit_Y2"
    ws_o["D4"] = "=Gross_Profit_Y3"

    path = FIXTURES_DIR / "gp_waterfall_compliant.xlsx"
    wb.save(path)
    return path


def create_violations() -> Path:
    """Create a GP Waterfall with intentional IDFA violations."""
    wb = openpyxl.Workbook()

    # --- Assumptions sheet (same as compliant) ---
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a["A1"] = "Input"
    ws_a["B1"] = "Value"

    _set_cell(ws_a, "A2", "Year 1 Revenue")
    _set_cell(ws_a, "B2", 10_000_000, comment="INTENT: Base year revenue\nGENERATED: fixture")
    _add_named_range(wb, "Inp_Rev_Y1", "Assumptions", "B2")

    _set_cell(ws_a, "A3", "Revenue Growth Rate")
    _set_cell(ws_a, "B3", 0.10, comment="INTENT: Annual revenue growth rate\nGENERATED: fixture")
    _add_named_range(wb, "Inp_Rev_Growth", "Assumptions", "B3")

    _set_cell(ws_a, "A4", "Year 1 COGS %")
    _set_cell(
        ws_a, "B4", 0.60, comment="INTENT: Year 1 COGS as percentage of revenue\nGENERATED: fixture"
    )
    _add_named_range(wb, "Inp_COGS_Pct_Y1", "Assumptions", "B4")

    _set_cell(ws_a, "A5", "Annual COGS Efficiency Gain")
    _set_cell(
        ws_a, "B5", 0.01, comment="INTENT: Annual COGS efficiency improvement\nGENERATED: fixture"
    )
    _add_named_range(wb, "Inp_COGS_Efficiency", "Assumptions", "B5")

    # --- Calculations sheet (with violations) ---
    ws_c = wb.create_sheet("Calculations")
    ws_c["A1"] = "Variable"
    ws_c["B1"] = "Formula"

    # Revenue_Y1 — compliant
    _set_cell(ws_c, "A2", "Revenue_Y1")
    _set_cell(
        ws_c,
        "B2",
        "=Inp_Rev_Y1",
        formula=True,
        comment="INTENT: Year 1 revenue equals the input assumption",
    )
    _add_named_range(wb, "Revenue_Y1", "Calculations", "B2")

    # VIOLATION: coordinate reference instead of Named Range
    _set_cell(ws_c, "A3", "Revenue_Y2")
    _set_cell(
        ws_c, "B3", "=B2*1.10", formula=True, comment="INTENT: Year 2 revenue = prior year * 1.10"
    )
    _add_named_range(wb, "Revenue_Y2", "Calculations", "B3")

    # Revenue_Y3 — compliant
    _set_cell(ws_c, "A4", "Revenue_Y3")
    _set_cell(
        ws_c,
        "B4",
        "=Revenue_Y2*(1+Inp_Rev_Growth)",
        formula=True,
        comment="INTENT: Year 3 revenue = prior year * (1 + growth)",
    )
    _add_named_range(wb, "Revenue_Y3", "Calculations", "B4")

    # COGS_Pct_Y1 — compliant
    _set_cell(ws_c, "A5", "COGS_Pct_Y1")
    _set_cell(
        ws_c,
        "B5",
        "=Inp_COGS_Pct_Y1",
        formula=True,
        comment="INTENT: Year 1 COGS % equals the input assumption",
    )
    _add_named_range(wb, "COGS_Pct_Y1", "Calculations", "B5")

    # COGS_Pct_Y2 — compliant
    _set_cell(ws_c, "A6", "COGS_Pct_Y2")
    _set_cell(
        ws_c,
        "B6",
        "=COGS_Pct_Y1-Inp_COGS_Efficiency",
        formula=True,
        comment="INTENT: Year 2 COGS % = prior year - efficiency gain",
    )
    _add_named_range(wb, "COGS_Pct_Y2", "Calculations", "B6")

    # COGS_Pct_Y3 — compliant
    _set_cell(ws_c, "A7", "COGS_Pct_Y3")
    _set_cell(
        ws_c,
        "B7",
        "=COGS_Pct_Y2-Inp_COGS_Efficiency",
        formula=True,
        comment="INTENT: Year 3 COGS % = prior year - efficiency gain",
    )
    _add_named_range(wb, "COGS_Pct_Y3", "Calculations", "B7")

    # VIOLATION: hardcoded constant in Calculation layer
    _set_cell(ws_c, "A8", "COGS_Y1")
    _set_cell(
        ws_c,
        "B8",
        "=Revenue_Y1*0.60",
        formula=True,
        comment="INTENT: Year 1 COGS = revenue * COGS %",
    )
    _add_named_range(wb, "COGS_Y1", "Calculations", "B8")

    # COGS_Y2 — compliant
    _set_cell(ws_c, "A9", "COGS_Y2")
    _set_cell(
        ws_c,
        "B9",
        "=Revenue_Y2*COGS_Pct_Y2",
        formula=True,
        comment="INTENT: Year 2 COGS = revenue * COGS %",
    )
    _add_named_range(wb, "COGS_Y2", "Calculations", "B9")

    # COGS_Y3 — compliant
    _set_cell(ws_c, "A10", "COGS_Y3")
    _set_cell(
        ws_c,
        "B10",
        "=Revenue_Y3*COGS_Pct_Y3",
        formula=True,
        comment="INTENT: Year 3 COGS = revenue * COGS %",
    )
    _add_named_range(wb, "COGS_Y3", "Calculations", "B10")

    # Gross_Profit_Y1 — compliant
    _set_cell(ws_c, "A11", "Gross_Profit_Y1")
    _set_cell(
        ws_c,
        "B11",
        "=Revenue_Y1-COGS_Y1",
        formula=True,
        comment="INTENT: Year 1 Gross Profit = revenue - COGS",
    )
    _add_named_range(wb, "Gross_Profit_Y1", "Calculations", "B11")

    # VIOLATION: missing Intent Note
    _set_cell(ws_c, "A12", "Gross_Profit_Y2")
    _set_cell(ws_c, "B12", "=Revenue_Y2-COGS_Y2", formula=True)  # No comment
    _add_named_range(wb, "Gross_Profit_Y2", "Calculations", "B12")

    # Gross_Profit_Y3 — compliant
    _set_cell(ws_c, "A13", "Gross_Profit_Y3")
    _set_cell(
        ws_c,
        "B13",
        "=Revenue_Y3-COGS_Y3",
        formula=True,
        comment="INTENT: Year 3 Gross Profit = revenue - COGS",
    )
    _add_named_range(wb, "Gross_Profit_Y3", "Calculations", "B13")

    # --- Output sheet ---
    ws_o = wb.create_sheet("Output")
    ws_o["A1"] = "Item"
    ws_o["B1"] = "Year 1"
    ws_o["C1"] = "Year 2"
    ws_o["D1"] = "Year 3"

    ws_o["A2"] = "Revenue"
    ws_o["B2"] = "=Revenue_Y1"
    ws_o["C2"] = "=Revenue_Y2"
    ws_o["D2"] = "=Revenue_Y3"

    ws_o["A3"] = "COGS"
    ws_o["B3"] = "=COGS_Y1"
    ws_o["C3"] = "=COGS_Y2"
    ws_o["D3"] = "=COGS_Y3"

    ws_o["A4"] = "Gross Profit"
    ws_o["B4"] = "=Gross_Profit_Y1"
    ws_o["C4"] = "=Gross_Profit_Y2"
    ws_o["D4"] = "=Gross_Profit_Y3"

    path = FIXTURES_DIR / "gp_waterfall_violations.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    p1 = create_compliant()
    print(f"Created: {p1}")
    p2 = create_violations()
    print(f"Created: {p2}")
    print("Done — 2 fixtures generated.")
